import openpyxl


def get_data(sheetName):

    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    main_list = []

    for rows in range(2, totalrows+1):
        datalist = []
        for cols in range(1, totalcols+1):
            data = sheet.cell(row=rows, column=cols).value
            print(data)
            datalist.insert(cols, data)
        main_list.insert(rows, datalist)
    return main_list

#def getdata_mainsheet(sheetName):






